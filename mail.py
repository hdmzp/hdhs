import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import threading
import pandas as pd
import numpy as np
import os
import re
import itertools
import webbrowser
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import matplotlib
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

matplotlib.rcParams['font.family'] = 'Malgun Gothic'
matplotlib.rcParams['axes.unicode_minus'] = False

_LINK_COUNTER = itertools.count()


# ──────────────────────────────────────────────
# 데이터 처리 로직
# ──────────────────────────────────────────────

def fmt_pct(val):
    if pd.isna(val):
        return "-"
    return f"{round(val)}%"


def sum_to_str(series):
    return round(series.sum() / 1e8, 1)


def calc_rate(df, nume_col, deno_col):
    # fillna(0): 목표누락PGM(순주문목표=0)의 실적값이 NaN이어도 집계에 포함
    nume = df[nume_col].fillna(0).sum()
    deno = df[deno_col].fillna(0).sum()
    return int(round(nume / deno * 100, 0)) if deno else 0


def contrib(df, teams, mode, subtract_fee=False):
    filtered_df = df[
        (df["사업부명"] == "식품사업부") &
        (df["팀명"].isin(teams)) &
        (df["순주문목표"] > 0) &
        (df["운영구분"] == mode)
    ]
    base = filtered_df["●순주문(생방)"].sum()
    if base == 0:
        return 0.0
    profit = filtered_df["●공헌(생방)"].sum()
    if subtract_fee:
        profit -= filtered_df["PGM수수료"].sum()
    return round(profit / base * 100, 0)


def count_time(df, day_type, hours):
    return sum((df["요일"].isin(day_type)) & (df["시간대"].isin(hours)))


def exclude_wang_rerun(df):
    """토요일 10·11시 왕영은의 톡 투게더 재방(순주문목표=0) 제외"""
    rerun_mask = (
        (df["PGM명"].astype(str).str.contains("왕영은의 톡 투게더", na=False)) &
        (df["순주문목표"] == 0) &
        (df["요일"] == "토") &
        (df["시간대"].isin([10, 11]))
    )
    return df[~rerun_mask]


FILE_DIR = r"C:/Users/Hhome/Downloads"
HL_RE = re.compile(r'(\[순[\d-]+%?\s*-\s*공[\d-]+%?\])')

# 협력사 통합 그룹 (키워드 포함 시 하나로 묶어 표시)
PARTNER_GROUPS = [
    {"keywords": ["에스엘바이오텍", "코스네이처"],  "display": "닥터린/뉴트리코어"},
    {"keywords": ["에스더포뮬러",   "큐어라벨"],    "display": "에스더포뮬러/큐어라벨"},
]

def normalize_partner(name):
    """협력사명 정규화: 그룹키워드 매칭 → 그룹명, 주식회사 → (주)"""
    s = str(name)
    for grp in PARTNER_GROUPS:
        if any(kw in s for kw in grp["keywords"]):
            return grp["display"]
    return s.replace("주식회사", "(주)")

# df_b는 파일에 따라 날짜 컬럼명이 다를 수 있어 후보를 모두 시도
DATE_COLS = ['기준일자', '접수일자', '방송일자', '일자', '날짜', '기준일', '방송날짜']


def find_latest_file(directory, pattern, target_date):
    matches = []
    for fname in os.listdir(directory):
        fpath = os.path.join(directory, fname)
        if os.path.isfile(fpath) and pattern in fname:
            ts = os.path.getctime(fpath)
            if datetime.fromtimestamp(ts).date() == target_date:
                matches.append((fpath, ts))
    if not matches:
        return None
    matches.sort(key=lambda x: x[1], reverse=True)
    return matches[0][0]


PARSE_FORMATS = [
    '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d',
    '%y-%m-%d', '%y/%m/%d', '%y.%m.%d', '%y%m%d',
    '%m/%d/%Y', '%d/%m/%Y',
]

def _try_parse(series):
    """여러 포맷을 순서대로 시도해 가장 많이 파싱되는 결과 반환"""
    s = series.astype(str).str.strip()
    # infer 먼저 시도
    p = pd.to_datetime(s, errors='coerce', infer_datetime_format=True)
    if p.notna().sum() > len(s) * 0.5:
        return p
    # 명시적 포맷 순회
    for fmt in PARSE_FORMATS:
        try:
            p = pd.to_datetime(s, format=fmt, errors='coerce')
            if p.notna().sum() > len(s) * 0.5:
                return p
        except Exception:
            continue
    return p  # 그래도 실패면 마지막 결과 반환


def filter_month(df, ref_ym):
    for col in DATE_COLS:
        if col in df.columns:
            sample = df[col].dropna().astype(str).iloc[:2].tolist()
            parsed = _try_parse(df[col])
            mask = parsed.dt.strftime("%Y-%m") == ref_ym
            filtered = df[mask].copy()
            return filtered, col, sample
    return df.copy(), None, []


def build_report(df_a, df_b):
    ref_date = datetime.today() - timedelta(days=1)
    today_str = ref_date.strftime("%y/%m/%d")
    ref_ym    = ref_date.strftime("%Y-%m")
    month_label = ref_date.strftime("%y년 %m월")

    # ★ 해당 월만 필터 — 사용된 날짜 컬럼명도 반환해서 진단에 표시
    df_a, dcol_a, _ = filter_month(df_a, ref_ym)
    df_b, dcol_b, samp_b = filter_month(df_b, ref_ym)

    diag_a = f"1차: {dcol_a or '날짜컬럼없음'} ({len(df_a):,}행)"
    diag_b = f"2차: {dcol_b or '날짜컬럼없음'} ({len(df_b):,}행)"
    # 0행이면 샘플값 표시해서 원인 파악
    diag_warn = ""
    if dcol_b and len(df_b) == 0:
        diag_warn = f"  ⚠ 2차 필터 결과 0행 — 날짜 샘플: {samp_b}  (형식 확인 필요)"
    elif not dcol_b:
        diag_warn = "  ⚠ 2차 파일 날짜 컬럼 미확인 — 누적 집계일 수 있음"

    teams = ["건강식품1팀", "건강식품2팀", "일반식품1팀", "일반식품2팀"]

    lines = []
    lines.append(("title",   f"📩 식품사업부 Live방송 Daily Summary"))
    lines.append(("sub",     f"  {month_label} 누계 기준  |  -{today_str} 기준"))
    lines.append(("sub",     f"  [{diag_a}  /  {diag_b}]"))
    if diag_warn:
        lines.append(("warn", diag_warn))
    lines.append(("gap", ""))

    for team in teams:
        lines.append(("team",    f"☑️  {team}"))
        df_team_b    = df_b[df_b['팀명'] == team]
        df_team_live = df_team_b[df_team_b['생방비생방'] == '생방송']
        df_team_re   = df_team_b[df_team_b['생방비생방'] == '非생방송']

        gmv_total    = sum_to_str(df_team_b['거래금액'] + df_team_b['방송수수료'] + df_team_b['기타광고비'])
        profit_total = sum_to_str(df_team_b['공헌이익금액'])
        adv_total    = sum_to_str(df_team_b['방송수수료'] + df_team_b['기타광고비'])
        lines.append(("normal", f"  ▸ [접수]  매출 {gmv_total}억 / 공헌 {profit_total}억 / 광고비 {adv_total}억"))

        gmv_live    = sum_to_str(df_team_live['거래금액'] + df_team_live['방송수수료'] + df_team_live['기타광고비'])
        profit_live = sum_to_str(df_team_live['공헌이익금액'])
        adv_live    = sum_to_str(df_team_live['방송수수료'] + df_team_live['기타광고비'])
        lines.append(("normal", f"  ▸ [방송]  매출 {gmv_live}억 / 공헌 {profit_live}억 / 광고비 {adv_live}억"))

        gmv_re    = sum_to_str(df_team_re['거래금액'] + df_team_re['방송수수료'])
        profit_re = sum_to_str(df_team_re['공헌이익금액'])
        lines.append(("normal", f"  ▸ [미주]  매출 {gmv_re}억 / 공헌 {profit_re}억"))

        df_team_a    = df_a[(df_a['팀명'] == team) & (df_a['시간대구분'] != '심야')]
        total_order  = calc_rate(df_team_a, '●순주문(전체)', '순주문목표')
        total_profit = calc_rate(df_team_a, '●공헌(전체)',   '공헌목표')
        live_order   = calc_rate(df_team_a, '●순주문(생방)', '순주문목표')
        live_profit  = calc_rate(df_team_a, '●공헌(생방)',   '공헌목표')

        lines.append(("section", "  💡 달성률 요약"))
        lines.append(("hl_line", f"    • 전체   달성률:  [순{total_order}% - 공{total_profit}%]"))
        lines.append(("hl_line", f"    • 생방송 달성률:  [순{live_order}% - 공{live_profit}%]"))

        df_gnl = df_a[
            (df_a['팀명'] == team) &
            (df_a["시간대구분"] != "심야") &
            df_a["PGM명"].notna()
        ]

        df_no_wang = df_gnl[~df_gnl["PGM명"].str.contains("왕영은", na=False)]
        순_분자 = df_no_wang["●순주문(전체)"].sum()
        순_분모 = df_no_wang["순주문목표"].sum()
        공_분자 = df_no_wang["●공헌(전체)"].sum()
        공_분모 = df_no_wang["공헌목표"].sum()
        if 순_분모 > 0 and not pd.isna(순_분자) and 공_분모 > 0 and not pd.isna(공_분자):
            순달1 = int(round(순_분자 / 순_분모 * 100, 0))
            공달1 = int(round(공_분자 / 공_분모 * 100, 0))
            lines.append(("hl_line", f"    • 왕톡 제외 달성률:  [순{fmt_pct(순달1)} - 공{fmt_pct(공달1)}]"))

        exclude_kw = ['왕영은', '황정민', '오감쇼', '메종']
        df_no_fixed = df_gnl[~df_gnl["PGM명"].str.contains('|'.join(exclude_kw), na=False)]
        순_분자2 = df_no_fixed["●순주문(전체)"].sum()
        순_분모2 = df_no_fixed["순주문목표"].sum()
        공_분자2 = df_no_fixed["●공헌(전체)"].sum()
        공_분모2 = df_no_fixed["공헌목표"].sum()
        if 순_분모2 > 0 and not pd.isna(순_분자2) and 공_분모2 > 0 and not pd.isna(공_분자2):
            순달2 = int(round(순_분자2 / 순_분모2 * 100, 0))
            공달2 = int(round(공_분자2 / 공_분모2 * 100, 0))
            lines.append(("hl_line", f"    • 고정PGM 제외 달성률:  [순{fmt_pct(순달2)} - 공{fmt_pct(공달2)}]"))

        pgm_map = {'왕톡': '왕영은', '황정민쇼': '황정민', '오감쇼': '오감쇼', '메종': '메종'}
        df_team_valid = df_a[df_a['팀명'] == team]
        has_가중분 = '가중분' in df_team_valid.columns
        pgm_lines = []
        for label, keyword in pgm_map.items():
            df_pgm = df_team_valid[df_team_valid['PGM명'].str.contains(keyword, na=False)]
            if has_가중분:
                df_pgm = df_pgm[df_pgm['가중분'] > 0]
            count = df_pgm['브랜드명'].nunique()
            if count > 0:
                brands = df_pgm.sort_values('기준일자')['브랜드명'].dropna().unique()
                pgm_lines.append(f"    • {label}: {count}회  ({', '.join(brands)})")

        lines.append(("section", "  💡 고정PGM 진행 현황"))
        if pgm_lines:
            for pl in pgm_lines:
                lines.append(("normal", pl))
        else:
            lines.append(("normal", "    진행이력 없음"))
        lines.append(("gap", ""))

    # ── 고효율 PGM TOP5
    lines.append(("divider", "─" * 56))
    lines.append(("section", "📌 고효율 PGM TOP5   ※ 생방송 달성률 기준"))
    for label, team_list in [("건강식품1팀", ["건강식품1팀"]),
                              ("건강식품2팀", ["건강식품2팀"]),
                              ("일반식품1,2팀", ["일반식품1팀", "일반식품2팀"])]:
        lines.append(("subhead", f"  🟢 {label}"))
        df_eff = df_a[
            (df_a["팀명"].isin(team_list)) &
            (df_a["시간대구분"] != "심야")
        ]
        df_eff = df_eff[~df_eff['PGM명'].astype(str).str.contains('|'.join(exclude_kw), na=False)]
        df_eff = exclude_wang_rerun(df_eff)
        for _, row in df_eff.sort_values('●순달(생방)', ascending=False).head(5).iterrows():
            d = pd.to_datetime(row['기준일자']).strftime("%m/%d")
            h = int(float(row["시간대"])) if pd.notna(row["시간대"]) else 0
            s = round(row["●순달(생방)"], 0)
            p = round(row["●공달(생방)"], 0)
            lines.append(("normal", f"    • {d}({row['요일']}) {h:02d}시  {row['브랜드명']}  순{s} / 공{p}"))

    # ── 매출 TOP10
    lines.append(("gap", ""))
    lines.append(("section", "📌 매출 TOP10   ※ 접수 거래금액 기준"))
    for label, team_list in [("건강식품1팀", ["건강식품1팀"]),
                              ("건강식품2팀", ["건강식품2팀"]),
                              ("일반식품1,2팀", ["일반식품1팀", "일반식품2팀"])]:
        lines.append(("subhead", f"  🟢 {label}"))
        top_pre = df_b[df_b["팀명"].isin(team_list)].groupby("브랜드명")["거래금액"].sum().sort_values(ascending=False).head(10)
        items = [f"{brand} {int(val / 1e8)}억" for brand, val in top_pre.items()]
        lines.append(("normal", f"    {',  '.join(items)}"))

    # ── PGM달성률 컬럼 계산 ((순달+공달)/2)
    valid_teams = ["건강식품1팀", "건강식품2팀", "일반식품1팀", "일반식품2팀"]
    df_base = df_a[
        (df_a["팀명"].isin(valid_teams)) &
        (df_a["시간대구분"] != "심야")
    ].dropna(subset=["시간대"])
    df_base = exclude_wang_rerun(df_base).copy()

    if '●공달(전체)' in df_base.columns:
        df_base['_pgm달성률'] = (df_base['●순달(전체)'].fillna(0) + df_base['●공달(전체)'].fillna(0)) / 2
    else:
        공달_calc = df_base.apply(
            lambda r: r['●공헌(전체)'] / r['공헌목표'] * 100 if pd.notna(r.get('공헌목표')) and r.get('공헌목표', 0) > 0 else 0,
            axis=1
        )
        df_base['_pgm달성률'] = (df_base['●순달(전체)'].fillna(0) + 공달_calc.fillna(0)) / 2

    df_low30 = df_base[df_base["●순달(전체)"] < 30.0]
    df_p50   = df_base[df_base["_pgm달성률"]  < 50.0]

    # ── 순주문 30% 미만 — 팀별 분리
    lines.append(("gap", ""))
    lines.append(("section", f"📌 순주문 30% 미만 PGM  ({len(df_low30)}개)   ※ 전체 순달 기준(미주 포함)"))
    for t in ["건강식품1팀", "건강식품2팀", "일반식품1팀", "일반식품2팀"]:
        df_t = df_low30[df_low30["팀명"] == t]
        if df_t.empty:
            continue
        lines.append(("subhead", f"  🟢 {t}  ({len(df_t)}개)"))
        for _, r in df_t.iterrows():
            d = pd.to_datetime(r["기준일자"]).strftime("%m/%d")
            h = int(float(r["시간대"]))
            lines.append(("warn", f"    ⚠ {d}({r['요일']}) {h:02d}시  {r['브랜드명']}  {round(r['●순달(전체)'], 0)}%"))

    # ── PGM달성률 50% 미만 — 팀별 분리
    lines.append(("gap", ""))
    lines.append(("section", f"📌 PGM달성률 50% 미만  ({len(df_p50)}개)   ※ (순달+공달)÷2 기준"))
    for t in ["건강식품1팀", "건강식품2팀", "일반식품1팀", "일반식품2팀"]:
        df_t = df_p50[df_p50["팀명"] == t]
        if df_t.empty:
            continue
        lines.append(("subhead", f"  🟢 {t}  ({len(df_t)}개)"))
        for _, r in df_t.iterrows():
            d = pd.to_datetime(r["기준일자"]).strftime("%m/%d")
            h = int(float(r["시간대"]))
            pgm달 = round(r["_pgm달성률"], 0)
            순달  = round(r["●순달(전체)"], 0)
            lines.append(("warn_mid", f"    • {d}({r['요일']}) {h:02d}시  {r['브랜드명']}  P{int(pgm달)}%  (순{int(순달)}%)"))

    # ── 건강식품 협력사별 달성률 — 팀별 분리
    partner_col = next((c for c in ['협력사명'] if c in df_a.columns), None)
    brand_col   = '브랜드명' if '브랜드명' in df_a.columns else None

    if partner_col:
        lines.append(("gap", ""))
        lines.append(("section", f"📌 건강식품 협력사별 달성률"))
        for t in ["건강식품1팀", "건강식품2팀"]:
            df_t = df_a[
                (df_a["팀명"] == t) &
                (df_a["시간대구분"] != "심야") &
                (df_a["순주문목표"] > 0)
            ].copy()
            if df_t.empty:
                continue
            # 협력사명 정규화 (그룹화 + 주식회사 치환)
            df_t["_partner"] = df_t[partner_col].apply(normalize_partner)
            lines.append(("subhead", f"  🟢 {t}"))
            lines.append(("table_hdr", "    협력사명  |  브랜드명  |  횟수  |  순달  |  공달"))
            rows_out = []
            for grp_name, grp in df_t.groupby("_partner"):
                cnt    = len(grp)
                s달    = calc_rate(grp, '●순주문(전체)', '순주문목표')
                c달    = calc_rate(grp, '●공헌(전체)',   '공헌목표')
                brands = ""
                if brand_col:
                    brand_list = grp[brand_col].dropna().unique().tolist()
                    brands = ", ".join(str(b) for b in brand_list)
                rows_out.append({"name": grp_name, "brands": brands, "count": cnt, "순달": s달, "공달": c달})
            rows_out.sort(key=lambda x: x["순달"], reverse=True)
            for row in rows_out:
                lines.append(("partner_row", row))
            lines.append(("gap", ""))

    # ── 특이사항
    df_missing = df_a[
        (df_a["사업부명"] == "식품사업부") &
        (df_a["시간대구분"] != "심야") &
        (df_a["순주문목표"] == 0) &
        (df_a["PGM명"] != "왕영은의 톡 투게더")
    ]
    missing_count = len(df_missing)
    missing_items = []
    for _, r in df_missing.iterrows():
        d = pd.to_datetime(r["기준일자"]).strftime("%m/%d")
        h = int(float(r["시간대"])) if pd.notna(r.get("시간대")) else 0
        b = r.get("브랜드명", "")
        missing_items.append(f"{d} {h:02d}시 {b}")
    missing_detail = ",  ".join(missing_items) if missing_items else ""

    lines.append(("gap", ""))
    lines.append(("section", "📌 특이사항"))
    if missing_detail:
        lines.append(("normal", f"  • 목표 누락 PGM {missing_count}개  ({missing_detail}  — 누락 시 달성률 오차 발생 가능)"))
    else:
        lines.append(("normal", f"  • 목표 누락 PGM {missing_count}개  (누락 시 달성률 오차 발생 가능)"))
    건_정률 = int(round(contrib(df_a, ['건강식품1팀','건강식품2팀'], '정률'), 0))
    건_정액 = int(round(contrib(df_a, ['건강식품1팀','건강식품2팀'], '정액', True), 0))
    일_정률 = int(round(contrib(df_a, ['일반식품1팀','일반식품2팀'], '정률'), 0))
    일_정액 = int(round(contrib(df_a, ['일반식품1팀','일반식품2팀'], '정액', True), 0))
    lines.append(("normal", f"  • 건강식품  정률 평균 공헌율 {건_정률}% / 정액 {건_정액}%"))
    lines.append(("normal", f"  • 일반식품  정률 평균 공헌율 {일_정률}% / 정액 {일_정액}%"))

    # ── 시간대별 운영 빈도
    df_health = df_a[df_a["팀명"].isin(["건강식품1팀", "건강식품2팀"])].copy()
    df_health["요일"] = df_health["요일"].astype(str)
    weekdays = ['월', '화', '수', '목', '금']
    weekends = ['토', '일']
    wdc = {h: count_time(df_health, weekdays, [h]) for h in [6, 15, 16, 1]}
    wec = {h: count_time(df_health, weekends, [h]) for h in [6, 7, 8, 1]}
    lines.append(("normal", f"  • 건강식품 시간대별 운영 빈도"))
    lines.append(("normal", f"    [평일] 06시 {wdc[6]}회 / 15시 {wdc[15]}회 / 16시 {wdc[16]}회 / 01시 {wdc[1]}회"))
    lines.append(("normal", f"    [주말] 06시 {wec[6]}회 / 07시 {wec[7]}회 / 08시 {wec[8]}회 / 01시 {wec[1]}회"))

    # ── 상생방송
    df_sangsaeng = df_a[
        (df_a["사업부명"] == "식품담당") &
        (df_a["행사타이틀"].astype(str).str.contains("상생", na=False))
    ]
    total_s   = len(df_sangsaeng)
    regular_s = len(df_sangsaeng[df_sangsaeng["시간대구분"] != "심야"])
    night_s   = len(df_sangsaeng[df_sangsaeng["시간대구분"] == "심야"])
    lines.append(("normal", f"  • 상생방송 {total_s}회  (정규 {regular_s}회 / 심야 {night_s}회)"))

    return lines


# ──────────────────────────────────────────────
# 오감쇼 데이터 처리 로직 (RPA(1차) 파일 기반)
# ──────────────────────────────────────────────

OGAMSHO_KEYWORDS = ['오감쇼']
KOREAN_WEEKDAYS = ['월', '화', '수', '목', '금', '토', '일']


def is_ogamsho(pgm_name):
    s = str(pgm_name)
    return any(k in s for k in OGAMSHO_KEYWORDS)


def eok(v):
    return f"{round(float(v or 0) / 1e8, 1):.1f}억"


def safe_pct(num, den):
    try:
        return int(round(float(num) / float(den) * 100)) if den and float(den) != 0 else 0
    except Exception:
        return 0


def fmt_time(t):
    t = int(float(t or 0))
    return f"{t // 100:02d}:{t % 100:02d}"


HMALL_LINK_TEMPLATE = "https://www.hmall.com/md/pda/itemPtc?slitmCd={code}&preview=true"


def fmt_code(v):
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def product_url(code):
    s = fmt_code(code)
    if not s or s.lower() == 'nan':
        return None
    return HMALL_LINK_TEMPLATE.format(code=s)


OGAMSHO_CAT_RULES = [
    ('명품', '잡화/주얼리'), ('이너웨어', '패션의류'), ('언더', '패션의류'),
    ('레포츠기타', '패션의류'), ('레포츠의류', '패션의류'), ('패션의류', '패션의류'),
    ('주방', '주방'), ('인테리어', '리빙'), ('생활/건강', '리빙'),
    ('다이슨', '가전'), ('듀얼소닉', '뷰티'), ('가전', '가전'),
    ('미용', '뷰티'), ('식품', '식품'), ('생활용품', '리빙'),
    ('패션잡화', '잡화/주얼리'),
]


def ogamsho_cat(s):
    s = str(s)
    for k, v in OGAMSHO_CAT_RULES:
        if k in s:
            return v
    return '기타'


def load_ogamsho_df(path):
    df = None
    for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
        try:
            df = pd.read_csv(path, encoding=enc)
            break
        except Exception:
            continue
    if df is None:
        raise ValueError("인코딩 오류로 파일을 읽을 수 없습니다.")

    df = df[df['PGM명'].apply(is_ogamsho)].copy()
    if df.empty:
        return df

    df['기준일자'] = pd.to_datetime(df['기준일자'])
    for c in ['순주문목표', '●순주문(전체)', '●순주문(생방)', '공헌목표', '●공헌(전체)', '●공헌(생방)',
              'PGM수수료', '노출분', '시작시각', '종료시각']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df['카테고리'] = df.get('상품군1차', '').apply(ogamsho_cat)
    return df


def agg_ogamsho(df, by):
    g = df.groupby(by, sort=False)
    r = g.agg(
        횟수=('기준일자', 'nunique'),
        매출=('●순주문(전체)', 'sum'),
        공헌=('●공헌(전체)', 'sum'),
        매출생=('●순주문(생방)', 'sum'),
        공헌생=('●공헌(생방)', 'sum'),
        광고비=('PGM수수료', 'sum'),
        순주문목표=('순주문목표', 'sum'),
        공헌목표=('공헌목표', 'sum'),
    ).reset_index()
    r['P']    = r['공헌'] - r['광고비']
    r['순달']  = r.apply(lambda x: safe_pct(x['매출'], x['순주문목표']), axis=1)
    r['공달']  = r.apply(lambda x: safe_pct(x['공헌'], x['공헌목표']), axis=1)
    r['순달생'] = r.apply(lambda x: safe_pct(x['매출생'], x['순주문목표']), axis=1)
    r['공달생'] = r.apply(lambda x: safe_pct(x['공헌생'], x['공헌목표']), axis=1)
    r['P달']  = r.apply(lambda x: safe_pct(x['P'],  x['공헌목표']), axis=1)
    return r


def ogamsho_monthly(df, year=None):
    d = df[df['기준일자'].dt.year == year].copy() if year else df.copy()
    if d.empty:
        return d
    d['YM'] = d['기준일자'].dt.to_period('M')
    r = agg_ogamsho(d, 'YM').sort_values('YM')
    r['기준년월'] = r['YM'].astype(str).apply(lambda s: s[:4] + "년 " + s[5:] + "월")
    return r


def build_ogamsho_report(df):
    if df.empty:
        return [("warn", "선택한 파일에서 오감쇼/황정민쇼 방송 데이터를 찾지 못했습니다.")]

    lines = []
    latest = df['기준일자'].max()
    rows = df[df['기준일자'] == latest].sort_values('시작시각')
    m, d = latest.month, latest.day
    wd = KOREAN_WEEKDAYS[latest.weekday()]

    lines.append(("title", f"📺 {m}/{d} 오감쇼 Daily Report"))
    lines.append(("sub",   f"  최신 방송일 기준  |  대상 PGM: {' / '.join(OGAMSHO_KEYWORDS)}"))
    lines.append(("gap", ""))

    # 당일 방송 실적
    ts, tst   = rows['●순주문(전체)'].sum(), rows['순주문목표'].sum()
    tc, tct   = rows['●공헌(전체)'].sum(), rows['공헌목표'].sum()
    ts_live   = rows['●순주문(생방)'].sum()
    tc_live   = rows['●공헌(생방)'].sum()
    ad_cost   = rows['PGM수수료'].sum()

    lines.append(("section", f"📌 {m}/{d}({wd}) 방송 실적"))
    lines.append(("normal",  f"  ▸ 매출 {eok(ts)} / 공헌 {eok(tc)} / 광고비 {eok(ad_cost)}"))
    lines.append(("hl_line", f"  ▸ 전체 달성률:  [순{safe_pct(ts,tst)}% - 공{safe_pct(tc,tct)}%]"))
    lines.append(("hl_line", f"  ▸ 생방송 달성률:  [순{safe_pct(ts_live,tst)}% - 공{safe_pct(tc_live,tct)}%]"))
    lines.append(("gap", ""))

    for _, r in rows.iterrows():
        brand   = str(r.get('브랜드명', ''))
        product = str(r.get('대표상품명', ''))
        url     = product_url(r.get('판매상품코드'))
        s, st = r['●순주문(전체)'], r['순주문목표']
        c, ct = r['●공헌(전체)'], r['공헌목표']
        s_live, c_live = r['●순주문(생방)'], r['●공헌(생방)']
        lines.append(("product_line", {"brand": brand, "product": product, "url": url}))
        lines.append(("hl_line", f"    ({fmt_time(r['시작시각'])}-{fmt_time(r['종료시각'])}, {int(r['노출분'])}분)  "
                                  f"[순{safe_pct(s,st)}% - 공{safe_pct(c,ct)}%]  "
                                  f"생방[순{safe_pct(s_live,st)}% - 공{safe_pct(c_live,ct)}%]"))
    lines.append(("gap", ""))

    # 누적 실적
    def cum(sub, lbl):
        매출  = sub['●순주문(전체)'].sum()
        공헌  = sub['●공헌(전체)'].sum()
        광고비 = sub['PGM수수료'].sum()
        return (f"  • {lbl}  {sub['기준일자'].nunique()}회  "
                f"매출 {eok(매출)} / 공헌 {eok(공헌)} / 광고비 {eok(광고비)}  "
                f"[순{safe_pct(매출, sub['순주문목표'].sum())}% - "
                f"공{safe_pct(공헌, sub['공헌목표'].sum())}%]")

    y       = latest.year
    yy      = str(y)[2:]
    y_prev  = y - 1
    yy_prev = str(y_prev)[2:]
    yr_prev = df[df['기준일자'].dt.year == y_prev]
    yr      = df[df['기준일자'].dt.year == y]
    h1      = yr[yr['기준일자'].dt.month.between(1, 3)]
    h2      = yr[yr['기준일자'].dt.month.between(4, 6)]

    lines.append(("section", "📌 누적 실적"))
    lines.append(("hl_line", cum(df, "전체")))
    lines.append(("hl_line", cum(yr_prev, f"{yy_prev}년")))
    lines.append(("hl_line", cum(yr, f"{yy}년 전체")))
    lines.append(("hl_line", cum(h1, f"{yy}년 1-3월")))
    lines.append(("hl_line", cum(h2, f"{yy}년 4-6월")))
    lines.append(("gap", ""))

    # 연도별 비교 표 (25년 vs 26년)
    def ysum(sub):
        s = sub['●순주문(전체)'].sum()
        c = sub['●공헌(전체)'].sum()
        st = sub['순주문목표'].sum()
        ct = sub['공헌목표'].sum()
        return s, safe_pct(s, st), safe_pct(c, ct)

    h1_prev = yr_prev[yr_prev['기준일자'].dt.month.between(1, 3)]
    h2_prev = yr_prev[yr_prev['기준일자'].dt.month.between(4, 6)]

    # 연간 누계는 전년 1/1~12/31 전체가 아니라 "올해와 같은 날짜까지"로 잘라 동기간으로 비교
    try:
        cutoff_prev = latest.replace(year=y_prev)
    except ValueError:
        cutoff_prev = latest.replace(year=y_prev, day=28)  # 2/29 등 윤년 예외 대응
    ytd_prev  = yr_prev[yr_prev['기준일자'] <= cutoff_prev]
    ytd_label = f"누계(1/1~{latest.month}/{latest.day})"

    lines.append(("section", f"📌 연도별 비교 ({yy_prev}년 vs {yy}년, 동기간 기준)"))
    lines.append(("ytable_hdr", ["구분", f"{yy_prev}년 매출", f"{yy}년 매출", "증감",
                                  f"{yy_prev}년 달성률", f"{yy}년 달성률"]))
    for label, sub_prev, sub_cur in [("1-3월", h1_prev, h1), ("4-6월", h2_prev, h2), (ytd_label, ytd_prev, yr)]:
        s_prev, sn_prev, sc_prev = ysum(sub_prev)
        s_cur,  sn_cur,  sc_cur  = ysum(sub_cur)
        if s_prev:
            delta = safe_pct(s_cur - s_prev, s_prev)
            delta_str = f"{'+' if delta >= 0 else ''}{delta}%"
        else:
            delta_str = "-"
        lines.append(("ytable_row", [label, eok(s_prev), eok(s_cur), delta_str,
                                      f"순{sn_prev}%-공{sc_prev}%", f"순{sn_cur}%-공{sc_cur}%"]))
    lines.append(("gap", ""))

    # 월별 운영 추이
    mon = ogamsho_monthly(df, latest.year)
    lines.append(("section", f"📌 {latest.year}년 월별 운영 추이"))
    if mon.empty:
        lines.append(("normal", "  데이터 없음"))
    else:
        for _, r in mon.iterrows():
            lines.append(("hl_line", f"  • {r['기준년월']}  {int(r['횟수'])}회  매출 {eok(r['매출'])} / 공헌 {eok(r['공헌'])}  "
                                      f"[순{r['순달']}% - 공{r['공달']}%]"))
        if len(mon) >= 2:
            t = mon.tail(2)
            prev, cur = t.iloc[-2], t.iloc[-1]
            delta = (cur['매출'] - prev['매출']) / max(prev['매출'], 1) * 100
            arrow = '▲' if delta > 0 else '▼'
            lines.append(("normal", f"  ▸ 전월대비: {arrow}{abs(delta):.0f}%  ({eok(prev['매출'])} → {eok(cur['매출'])})"))
        latest_row = mon.iloc[-1]
        lines.append(("normal", f"  ▸ 최근월: {latest_row['기준년월']} (순{latest_row['순달']}% / 공{latest_row['공달']}%)"))
    lines.append(("gap", ""))

    # 월별 카테고리 비중
    lines.append(("section", f"📌 {latest.year}년 월별 카테고리 비중"))
    if yr.empty:
        lines.append(("normal", "  데이터 없음"))
    else:
        yr_ym = yr.copy()
        yr_ym['YM'] = yr_ym['기준일자'].dt.to_period('M')
        for ym, grp in yr_ym.groupby('YM'):
            cat_g = agg_ogamsho(grp, '카테고리')
            tot = cat_g['순주문목표'].sum()
            cat_g['비중'] = cat_g['순주문목표'].apply(lambda x: safe_pct(x, tot))
            cat_g = cat_g.sort_values('비중', ascending=False)
            label = f"{str(ym)[:4]}년 {str(ym)[5:]}월"
            parts = ' · '.join(f"{r['카테고리']} {r['비중']}% ({int(r['횟수'])}회)" for _, r in cat_g.iterrows())
            lines.append(("breakdown", f"  • {label}   {parts}"))
    lines.append(("gap", ""))

    # 카테고리 Top
    if not yr.empty:
        lines.append(("section", f"📌 {latest.year}년 카테고리별 편성비중"))
        cat_w = agg_ogamsho(yr, '카테고리')
        total_target = cat_w['순주문목표'].sum()
        cat_w['비중'] = cat_w['순주문목표'].apply(lambda x: safe_pct(x, total_target))
        cat_w = cat_w.sort_values('비중', ascending=False)
        for _, r in cat_w.iterrows():
            lines.append(("normal", f"  • {r['카테고리']}  {r['비중']}%  ({int(r['횟수'])}회)"))
            cat_brands = yr[yr['카테고리'] == r['카테고리']]['브랜드명'].dropna().unique()
            if len(cat_brands):
                lines.append(("breakdown", f"      - {', '.join(cat_brands)}"))
        lines.append(("gap", ""))

        lines.append(("section", f"📌 {latest.year}년 카테고리별 순위 (매출순)"))
        rank_cat = agg_ogamsho(yr, '카테고리').sort_values('매출', ascending=False).reset_index(drop=True)
        for i, r in rank_cat.iterrows():
            lines.append(("hl_line", f"  {i + 1}. {r['카테고리']}  매출 {eok(r['매출'])}  공헌 {eok(r['공헌'])}  "
                                      f"[순{r['순달']}% - 공{r['공달']}%]  생방[순{r['순달생']}% - 공{r['공달생']}%]"))
        lines.append(("gap", ""))

    return lines


def build_ogamsho_tables(df):
    """GUI 표/그래프 렌더링용 구조화 데이터 (build_ogamsho_report와 별도, 같은 집계 로직 재사용)"""
    if df.empty:
        return None

    latest = df['기준일자'].max()
    rows = df[df['기준일자'] == latest].sort_values('시작시각')
    m, d = latest.month, latest.day
    wd = KOREAN_WEEKDAYS[latest.weekday()]

    daily_rows = []
    for _, r in rows.iterrows():
        s, st = r['●순주문(전체)'], r['순주문목표']
        c, ct = r['●공헌(전체)'], r['공헌목표']
        s_live, c_live = r['●순주문(생방)'], r['●공헌(생방)']
        daily_rows.append({
            "brand": str(r.get('브랜드명', '')),
            "product": str(r.get('대표상품명', '')),
            "url": product_url(r.get('판매상품코드')),
            "time": f"{fmt_time(r['시작시각'])}-{fmt_time(r['종료시각'])}",
            "expo": int(r['노출분']),
            "sn": safe_pct(s, st), "sc": safe_pct(c, ct),
            "ln": safe_pct(s_live, st), "lc": safe_pct(c_live, ct),
        })

    ts, tst = rows['●순주문(전체)'].sum(), rows['순주문목표'].sum()
    tc, tct = rows['●공헌(전체)'].sum(), rows['공헌목표'].sum()
    ts_live, tc_live = rows['●순주문(생방)'].sum(), rows['●공헌(생방)'].sum()
    ad_cost = rows['PGM수수료'].sum()
    daily_summary = {
        "label": f"{m}/{d}({wd})", "sales": eok(ts), "profit": eok(tc), "ad": eok(ad_cost),
        "sn": safe_pct(ts, tst), "sc": safe_pct(tc, tct),
        "ln": safe_pct(ts_live, tst), "lc": safe_pct(tc_live, tct),
    }

    def cum_row(sub, label):
        s, c, ad = sub['●순주문(전체)'].sum(), sub['●공헌(전체)'].sum(), sub['PGM수수료'].sum()
        st, ct = sub['순주문목표'].sum(), sub['공헌목표'].sum()
        return {"label": label, "count": int(sub['기준일자'].nunique()),
                "sales": eok(s), "profit": eok(c), "ad": eok(ad),
                "sn": safe_pct(s, st), "sc": safe_pct(c, ct)}

    y, yy = latest.year, str(latest.year)[2:]
    y_prev, yy_prev = y - 1, str(y - 1)[2:]
    yr_prev = df[df['기준일자'].dt.year == y_prev]
    yr      = df[df['기준일자'].dt.year == y]
    h1      = yr[yr['기준일자'].dt.month.between(1, 3)]
    h2      = yr[yr['기준일자'].dt.month.between(4, 6)]

    cumulative = [
        cum_row(df, "전체"),
        cum_row(yr_prev, f"{yy_prev}년"),
        cum_row(yr, f"{yy}년 전체"),
        cum_row(h1, f"{yy}년 1-3월"),
        cum_row(h2, f"{yy}년 4-6월"),
    ]

    h1_prev = yr_prev[yr_prev['기준일자'].dt.month.between(1, 3)]
    h2_prev = yr_prev[yr_prev['기준일자'].dt.month.between(4, 6)]
    try:
        cutoff_prev = latest.replace(year=y_prev)
    except ValueError:
        cutoff_prev = latest.replace(year=y_prev, day=28)
    ytd_prev  = yr_prev[yr_prev['기준일자'] <= cutoff_prev]
    ytd_label = f"누계(1/1~{m}/{d})"

    def ysum(sub):
        s, c = sub['●순주문(전체)'].sum(), sub['●공헌(전체)'].sum()
        st, ct = sub['순주문목표'].sum(), sub['공헌목표'].sum()
        return s, safe_pct(s, st), safe_pct(c, ct)

    yoy = []
    for label, sub_prev, sub_cur in [("1-3월", h1_prev, h1), ("4-6월", h2_prev, h2), (ytd_label, ytd_prev, yr)]:
        s_prev, sn_prev, sc_prev = ysum(sub_prev)
        s_cur,  sn_cur,  sc_cur  = ysum(sub_cur)
        delta = safe_pct(s_cur - s_prev, s_prev) if s_prev else None
        yoy.append({"label": label, "sales_prev": eok(s_prev), "sales_cur": eok(s_cur),
                     "delta": delta, "sn_prev": sn_prev, "sc_prev": sc_prev,
                     "sn_cur": sn_cur, "sc_cur": sc_cur})

    mon = ogamsho_monthly(df, y)
    monthly, highlight_ym = [], None
    if not mon.empty:
        highlight_ym = mon.iloc[-1]['기준년월']  # 가장 최근월 강조
        for _, r in mon.iterrows():
            monthly.append({"label": r['기준년월'], "count": int(r['횟수']),
                             "sales": eok(r['매출']), "profit": eok(r['공헌']),
                             "sn": int(r['순달']), "sc": int(r['공달'])})
        tot_s, tot_c = yr['●순주문(전체)'].sum(), yr['●공헌(전체)'].sum()
        tot_st, tot_ct = yr['순주문목표'].sum(), yr['공헌목표'].sum()
        monthly.append({"label": f"{yy}년 총계", "count": int(yr['기준일자'].nunique()),
                         "sales": eok(tot_s), "profit": eok(tot_c),
                         "sn": safe_pct(tot_s, tot_st), "sc": safe_pct(tot_c, tot_ct),
                         "is_total": True})

    category = []
    if not yr.empty:
        cat = agg_ogamsho(yr, '카테고리')
        total_target = cat['순주문목표'].sum()
        cat['비중'] = cat['순주문목표'].apply(lambda x: safe_pct(x, total_target))
        cat = cat.sort_values('매출', ascending=False)
        for _, r in cat.iterrows():
            category.append({"label": r['카테고리'], "count": int(r['횟수']), "weight": int(r['비중']),
                              "sales": eok(r['매출']), "profit": eok(r['공헌']),
                              "sn": int(r['순달']), "sc": int(r['공달']),
                              "sn_live": int(r['순달생']), "sc_live": int(r['공달생']),
                              "weight_raw": max(r['비중'], 0.1)})
        tot_s, tot_c = yr['●순주문(전체)'].sum(), yr['●공헌(전체)'].sum()
        tot_st, tot_ct = yr['순주문목표'].sum(), yr['공헌목표'].sum()
        tot_sl, tot_cl = yr['●순주문(생방)'].sum(), yr['●공헌(생방)'].sum()
        category.append({"label": "Total", "count": int(yr['기준일자'].nunique()), "weight": 100,
                          "sales": eok(tot_s), "profit": eok(tot_c),
                          "sn": safe_pct(tot_s, tot_st), "sc": safe_pct(tot_c, tot_ct),
                          "sn_live": safe_pct(tot_sl, tot_st), "sc_live": safe_pct(tot_cl, tot_ct),
                          "weight_raw": 0, "is_total": True})

    # 신상품/기존상품 비교 (연도별) — '신/첫' 컬럼이 있는 RPA(1차) 파일만 지원
    newproduct = []
    if '신/첫' in df.columns:
        def np_part(part_df, label, denom_target, is_total=False, is_new=False):
            s, c = part_df['●순주문(전체)'].sum(), part_df['●공헌(전체)'].sum()
            st, ct = part_df['순주문목표'].sum(), part_df['공헌목표'].sum()
            cnt = len(part_df)
            return {"label": label, "count": cnt, "weight": safe_pct(st, denom_target),
                    "sn": safe_pct(s, st), "sc": safe_pct(c, ct),
                    "is_total": is_total, "is_new": is_new}

        np_groups = [
            (f"{yy_prev}년", yr_prev),
            (f"{yy}년", yr),
            ("누계", df[df['기준일자'].dt.year.isin([y_prev, y])]),
        ]
        for year_label, sub in np_groups:
            # 순주문목표가 0/없는 방송(목표 미설정)은 신상품·기존상품 카운트에서 제외
            sub = sub[sub['순주문목표'] > 0]
            denom_target = sub['순주문목표'].sum()  # 비중(%) = 순주문목표 기준 (카테고리 비중과 동일 기준)
            new = sub[sub['신/첫'].isin(['신상품', '첫노출'])]
            old = sub[~sub['신/첫'].isin(['신상품', '첫노출'])]  # 기존상품 = 신상품/첫노출이 아닌 상품
            for row in (np_part(old, "기존상품", denom_target),
                        np_part(new, "신상품/첫노출", denom_target, is_new=True),
                        np_part(sub, "총계", denom_target, is_total=True)):
                row["year"] = year_label
                newproduct.append(row)

    return {
        "title": f"{m}/{d} 오감쇼 Daily Report", "year": y,
        "daily_summary": daily_summary, "daily_rows": daily_rows,
        "cumulative": cumulative, "yoy": yoy, "monthly": monthly, "highlight_ym": highlight_ym,
        "category": category, "newproduct": newproduct,
    }


def quarter_range(month):
    q_start = ((month - 1) // 3) * 3 + 1
    return q_start, q_start + 2


def product_history(df, code, brand, before_date):
    """동일 상품(코드 우선, 없으면 브랜드)의 과거(이전 날짜) 진행 이력 집계"""
    s = fmt_code(code) if code is not None else ''
    if s and s.lower() != 'nan':
        hist = df[(df['판매상품코드'].apply(fmt_code) == s) & (df['기준일자'] < before_date)]
    elif brand:
        hist = df[(df['브랜드명'] == brand) & (df['기준일자'] < before_date)]
    else:
        return None
    if hist.empty:
        return None
    cnt = hist['기준일자'].nunique()
    sale = hist['●순주문(전체)'].sum()
    target = hist['순주문목표'].sum()
    return cnt, safe_pct(sale, target)


def build_ogamsho_comment(df):
    """가장 최근 방송일 기준 '보고멘트' 초안 생성 (실적 수치는 자동, 리뷰 문구는 직접 작성)"""
    if df.empty:
        return "선택한 파일에서 오감쇼 방송 데이터를 찾지 못했습니다."

    latest = df['기준일자'].max()
    rows = df[df['기준일자'] == latest].sort_values('시작시각')
    wd = KOREAN_WEEKDAYS[latest.weekday()]
    date_label = f"{latest.month}/{latest.day} {wd}"

    L = []
    L.append(f"■ {date_label} 오감쇼")
    L.append("ㆍ합계")
    ts, tst = rows['●순주문(전체)'].sum(), rows['순주문목표'].sum()
    tc, tct = rows['●공헌(전체)'].sum(), rows['공헌목표'].sum()
    start, end = rows['시작시각'].min(), rows['종료시각'].max()
    total_expo = int(rows['노출분'].sum())
    L.append(f"  ({fmt_time(start)}-{fmt_time(end)}, {total_expo}분)")
    L.append(f"  순주문 {eok(ts)}, {safe_pct(ts, tst)}%")
    L.append(f"  공헌 {eok(tc)}, {safe_pct(tc, tct)}%")
    L.append("")

    for _, r in rows.iterrows():
        brand   = str(r.get('브랜드명', '')).strip()
        product = str(r.get('대표상품명', '')).strip()
        name = f"{brand} {product}".strip()
        s, st = r['●순주문(전체)'], r['순주문목표']
        c, ct = r['●공헌(전체)'], r['공헌목표']
        L.append(f"ㆍ{name}")
        L.append(f"  ({fmt_time(r['시작시각'])}-{fmt_time(r['종료시각'])}, {int(r['노출분'])}분)")
        L.append(f"  순주문 {eok(s)}, {safe_pct(s, st)}%")
        L.append(f"  공헌 {eok(c)}, {safe_pct(c, ct)}%")
        L.append("")

    L.append("■ 누적 실적")

    def cum_simple(sub, label):
        s, c = sub['●순주문(전체)'].sum(), sub['●공헌(전체)'].sum()
        st, ct = sub['순주문목표'].sum(), sub['공헌목표'].sum()
        return f"{label} {sub['기준일자'].nunique()}회, 순{safe_pct(s, st)}%- 공{safe_pct(c, ct)}%"

    y = latest.year
    yr = df[df['기준일자'].dt.year == y]
    q_start, q_end = quarter_range(latest.month)
    qr = yr[yr['기준일자'].dt.month.between(q_start, q_end)]

    L.append(cum_simple(df, "전체"))
    L.append(cum_simple(yr, "금년"))
    L.append(cum_simple(qr, f"{q_start}-{q_end}월"))
    L.append("")

    L.append("■ 방송 리뷰")
    for _, r in rows.iterrows():
        brand   = str(r.get('브랜드명', '')).strip()
        product = str(r.get('대표상품명', '')).strip()
        name = f"{brand} {product}".strip()
        L.append(f"ㆍ{name}")
        hist = product_history(df, r.get('판매상품코드'), r.get('브랜드명'), latest)
        if hist:
            cnt, avg = hist
            L.append(f"  ({cnt + 1}회차, 직전 {cnt}회 평균 순{avg}%)")
        else:
            L.append("  (신규/첫 진행)")
        L.append("  (리뷰 내용을 입력해주세요)")
        L.append("")

    L.append("ㆍ차주는 아이템A, 아이템B, 아이템C 진행 예정입니다.")

    return "\n".join(L)


def create_ogamsho_excel(df, report_lines):
    today   = datetime.now().strftime('%Y%m%d')
    outpath = os.path.join(FILE_DIR, f"오감쇼_분석_{today}.xlsx")

    wb          = openpyxl.Workbook()
    fill_blue   = PatternFill("solid", fgColor="4472C4")
    fill_purple = PatternFill("solid", fgColor="7030A0")
    fill_green  = PatternFill("solid", fgColor="70AD47")
    white_bold  = Font(bold=True, color="FFFFFF")
    bold        = Font(bold=True)
    center      = Alignment(horizontal='center', vertical='center')

    def write_table(ws, start_row, headers, rows_data, fill):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=start_row, column=ci, value=h)
            c.fill, c.font, c.alignment = fill, white_bold, center
        for ri, row in enumerate(rows_data, start_row + 1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(horizontal='center')
        return start_row + 1 + len(rows_data)

    # 시트1: 보고멘트
    ws1 = wb.active
    ws1.title = "보고멘트"
    ws1.sheet_properties.tabColor = "FFFF00"
    ws1.column_dimensions['A'].width = 90
    for i, (tag, text) in enumerate(report_lines, 1):
        if not isinstance(text, str):
            continue
        c = ws1.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True)
        if tag == "title":
            c.font = Font(bold=True, size=13)
        elif tag == "section":
            c.font = bold

    # 시트2: 월별 운영현황
    ws2 = wb.create_sheet("월별 운영현황")
    ws2.sheet_properties.tabColor = "CC99FF"
    mon = ogamsho_monthly(df, datetime.now().year)
    MON_HDR  = ['기준년월', '횟수', '매출(억)', '공헌(억)', '광고비(억)', '순달(%)', '공달(%)', 'P달(%)']
    mon_rows = []
    for _, r in mon.iterrows():
        mon_rows.append([r['기준년월'], int(r['횟수']),
                          round(r['매출'] / 1e8, 1), round(r['공헌'] / 1e8, 1), round(r['광고비'] / 1e8, 1),
                          int(r['순달']), int(r['공달']), int(r['P달'])])
    write_table(ws2, 1, MON_HDR, mon_rows, fill_purple)
    for col in 'ABCDEFGH':
        ws2.column_dimensions[col].width = 14

    # 시트3: 카테고리별 결과
    ws3 = wb.create_sheet("카테고리별 결과")
    CAT_HDR  = ['카테고리', '편성비중(%)', '횟수', '매출(억)', '공헌(억)', '광고비(억)', '순달(%)', '공달(%)', 'P달(%)']
    cat_agg  = agg_ogamsho(df, '카테고리')
    total_target = cat_agg['순주문목표'].sum()
    cat_agg['비중'] = cat_agg['순주문목표'].apply(lambda x: safe_pct(x, total_target))
    cat_agg = cat_agg.sort_values('매출', ascending=False)
    cat_rows = []
    for _, r in cat_agg.iterrows():
        cat_rows.append([r['카테고리'], int(r['비중']), int(r['횟수']),
                          round(r['매출'] / 1e8, 1), round(r['공헌'] / 1e8, 1), round(r['광고비'] / 1e8, 1),
                          int(r['순달']), int(r['공달']), int(r['P달'])])
    write_table(ws3, 1, CAT_HDR, cat_rows, fill_green)
    for col in 'ABCDEFGHI':
        ws3.column_dimensions[col].width = 14

    # 시트4: RAWDATA
    ws4  = wb.create_sheet("RAWDATA")
    cols = list(df.columns)
    for ci, h in enumerate(cols, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = fill_blue, white_bold, center
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            v = row[col]
            if isinstance(v, pd.Timestamp):
                v = v.strftime('%Y-%m-%d')
            elif isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = float(v)
            ws4.cell(row=ri, column=ci, value=v)

    wb.save(outpath)
    return outpath


# ──────────────────────────────────────────────
# 텍스트 위젯 렌더링
# ──────────────────────────────────────────────

def setup_tags(box):
    box.tag_configure("title",      font=("맑은 고딕", 14, "bold"), foreground="#0d2550",  spacing1=6,  spacing3=12)
    box.tag_configure("sub",        font=("맑은 고딕",  8),         foreground="#9098a3",  spacing3=10)
    box.tag_configure("team",       font=("맑은 고딕", 10, "bold"), foreground="#ffffff",  background="#3a6cf4", spacing1=12, spacing3=8, lmargin1=6, lmargin2=6)
    box.tag_configure("section",      font=("맑은 고딕",  9, "bold"), foreground="#1a3a6e",  spacing1=12, spacing3=5)
    box.tag_configure("section_box",  font=("맑은 고딕",  9, "bold"), foreground="#ffffff",  background="#26314a", spacing1=14, spacing3=8, lmargin1=6, lmargin2=6)
    box.tag_configure("subhead",      font=("맑은 고딕",  9, "bold"), foreground="#1f9d55",  spacing1=6,  spacing3=3)
    box.tag_configure("normal",       font=("맑은 고딕",  9),         foreground="#3a3f47",  spacing1=2,  spacing3=2,  spacing2=3)
    box.tag_configure("hl_base",      font=("맑은 고딕",  9),         foreground="#3a3f47",  spacing1=2,  spacing3=4,  spacing2=3)
    box.tag_configure("hl",           font=("맑은 고딕",  9, "bold"), foreground="#0860d6",  background="#e6f2ff", spacing1=2, spacing3=4)
    box.tag_configure("breakdown",    font=("맑은 고딕",  8),         foreground="#6b7280",  spacing1=2,  spacing3=2)
    box.tag_configure("warn",         font=("맑은 고딕",  9, "bold"), foreground="#d6311a",  spacing1=2,  spacing3=2)
    box.tag_configure("warn_mid",     font=("맑은 고딕",  9, "bold"), foreground="#d97c00",  spacing1=2,  spacing3=2)
    box.tag_configure("divider",      font=("맑은 고딕",  8),         foreground="#dde1e7",  spacing1=8,  spacing3=4)
    box.tag_configure("gap",          font=("맑은 고딕",  6))
    box.tag_configure("table_hdr",    font=("맑은 고딕",  8),         foreground="#9098a3",  spacing1=5,  spacing3=3)
    box.tag_configure("t_name",       font=("맑은 고딕",  9, "bold"), foreground="#1a1a2e")
    box.tag_configure("t_공달",       font=("맑은 고딕",  9, "bold"), foreground="#d6311a")
    box.tag_configure("t_공달_low",   font=("맑은 고딕",  9, "bold"), foreground="#d6311a",  underline=True)
    box.tag_configure("t_normal",     font=("맑은 고딕",  9),         foreground="#6b7280")
    box.tag_configure("tbl_hdr",      font=("Consolas",  9, "bold"), foreground="#ffffff",  background="#3a6cf4",
                       tabs=(90, 190, 290, 360, 480, 600), spacing1=10, spacing3=4, lmargin1=4)
    box.tag_configure("tbl_row",      font=("Consolas",  9),         foreground="#333333",
                       tabs=(90, 190, 290, 360, 480, 600), spacing1=3,  spacing3=3, lmargin1=4)


def insert_line(box, tag, text):
    if tag == "hl_line":
        parts = HL_RE.split(text)
        for part in parts:
            if HL_RE.match(part):
                box.insert("end", part, "hl")
            else:
                box.insert("end", part, "hl_base")
        box.insert("end", "\n")
    elif tag == "gap":
        box.insert("end", "\n", "gap")
    elif tag == "partner_row":
        d = text
        공달_tag = "t_공달_low" if d['공달'] < 85 else "t_공달"
        box.insert("end", "    • ")
        box.insert("end", f"{d['name']}", "t_name")
        if d.get("brands"):
            box.insert("end", f"  ({d['brands']})", "t_normal")
        box.insert("end", f"  {d['count']}회", "t_normal")
        box.insert("end", f"  순{d['순달']}%", "t_normal")
        box.insert("end", f"  공{d['공달']}%\n", 공달_tag)
    elif tag == "section":
        # 📌 헤더는 자동으로 검정 박스 스타일 적용
        actual_tag = "section_box" if text.lstrip().startswith("📌") else "section"
        box.insert("end", text + "\n", actual_tag)
    elif tag in ("ytable_hdr", "ytable_row"):
        box.insert("end", "\t".join(text) + "\n", "tbl_hdr" if tag == "ytable_hdr" else "tbl_row")
    elif tag == "product_line":
        d = text
        box.insert("end", f"  • {d['brand']}  ", "normal")
        if d.get("url"):
            link_tag = f"link_{next(_LINK_COUNTER)}"
            box.tag_configure(link_tag, foreground="#0860d6", underline=True)
            box.tag_bind(link_tag, "<Enter>", lambda e, b=box: b.configure(cursor="hand2"))
            box.tag_bind(link_tag, "<Leave>", lambda e, b=box: b.configure(cursor=""))
            box.tag_bind(link_tag, "<Button-1>", lambda e, u=d['url']: webbrowser.open(u))
            box.insert("end", d['product'], ("normal", link_tag))
        else:
            box.insert("end", d['product'], "normal")
        box.insert("end", "\n")
    else:
        box.insert("end", text + "\n", tag)


# ──────────────────────────────────────────────
# GUI
# ──────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("RPA 현황 알리미")
        self.configure(bg="#eef1f6")
        self.minsize(780, 660)

        self.style = ttk.Style(self)
        try:
            self.style.theme_use('clam')
        except Exception:
            pass

        self._base_size = 9   # 모든 탭이 공유하는 글자 크기

        # ttk.Notebook의 "선택된 탭이 더 커 보이는" 테마 동작을 피하기 위해
        # 일반 tk.Button으로 직접 탭 바를 구현 (모든 버튼은 항상 동일한 크기)
        container = tk.Frame(self, bg="#eef1f6")
        container.pack(fill="both", expand=True, padx=10, pady=10)

        tab_bar = tk.Frame(container, bg="#eef1f6")
        tab_bar.pack(fill="x", side="top")

        body = tk.Frame(container, bg="#eef1f6")
        body.pack(fill="both", expand=True, pady=(8, 0))

        food_tab    = tk.Frame(body, bg="#eef1f6")
        og_tab      = tk.Frame(body, bg="#eef1f6")
        comment_tab = tk.Frame(body, bg="#eef1f6")
        food_tab.grid(row=0, column=0, sticky="nsew")
        og_tab.grid(row=0, column=0, sticky="nsew")
        comment_tab.grid(row=0, column=0, sticky="nsew")
        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)

        self._tab_btn = {}
        self._tab_frame = {"food": food_tab, "ogamsho": og_tab, "comment": comment_tab}

        def make_tab_button(key, text):
            btn = tk.Button(tab_bar, text=text, font=("맑은 고딕", 10, "bold"),
                             bd=0, relief="flat", padx=18, pady=10,
                             command=lambda: self._show_tab(key))
            btn.pack(side="left", padx=(0, 4))
            self._tab_btn[key] = btn

        make_tab_button("food", "식품사업부 Daily")
        make_tab_button("ogamsho", "오감쇼 Daily")
        make_tab_button("comment", "보고멘트")

        self._build_food_tab(food_tab)
        self._build_ogamsho_tab(og_tab)
        self._build_comment_tab(comment_tab)

        self._show_tab("food")

    def _show_tab(self, key):
        for k, btn in self._tab_btn.items():
            active = (k == key)
            btn.configure(background="#3a6cf4" if active else "#dde2eb",
                          activebackground="#3a6cf4" if active else "#dde2eb",
                          foreground="#ffffff" if active else "#5a6472",
                          activeforeground="#ffffff" if active else "#5a6472")
        self._tab_frame[key].tkraise()

    # ── 식품사업부 탭 ─────────────────────────
    def _build_food_tab(self, root):
        pad = {"padx": 12, "pady": 7}

        tk.Label(root, text="📩 식품사업부 Live방송 Daily 알리미",
                 font=("맑은 고딕", 15, "bold"), bg="#eef1f6", fg="#10254f").grid(
            row=0, column=0, columnspan=3, pady=(18, 10))

        tk.Label(root, text="1차 파일", bg="#eef1f6", font=("맑은 고딕", 9)).grid(row=1, column=0, sticky="e", **pad)
        self.path_a = tk.StringVar()
        tk.Entry(root, textvariable=self.path_a, width=54, state="readonly").grid(row=1, column=1, sticky="ew", **pad)
        tk.Button(root, text="찾아보기", command=self._pick_a, width=8).grid(row=1, column=2, **pad)

        tk.Label(root, text="2차 파일", bg="#eef1f6", font=("맑은 고딕", 9)).grid(row=2, column=0, sticky="e", **pad)
        self.path_b = tk.StringVar()
        tk.Entry(root, textvariable=self.path_b, width=54, state="readonly").grid(row=2, column=1, sticky="ew", **pad)
        tk.Button(root, text="찾아보기", command=self._pick_b, width=8).grid(row=2, column=2, **pad)

        btn_frame = tk.Frame(root, bg="#eef1f6")
        btn_frame.grid(row=3, column=0, columnspan=3, pady=(8, 6))

        self.btn_run = tk.Button(btn_frame, text="  보고서 생성  ", font=("맑은 고딕", 10, "bold"),
                                 bg="#3a6cf4", fg="white", activebackground="#2756d1",
                                 bd=0, padx=16, pady=9, command=self._run)
        self.btn_run.pack(side="left", padx=6)

        self.btn_copy = tk.Button(btn_frame, text="  클립보드 복사  ", font=("맑은 고딕", 10),
                                  bg="#22a85a", fg="white", activebackground="#1a8247",
                                  bd=0, padx=16, pady=9, command=self._copy, state="disabled")
        self.btn_copy.pack(side="left", padx=6)

        self.btn_clear = tk.Button(btn_frame, text="  지우기  ", font=("맑은 고딕", 10),
                                   bg="#888", fg="white", activebackground="#666",
                                   bd=0, padx=16, pady=9, command=self._clear)
        self.btn_clear.pack(side="left", padx=6)

        # 글자 크기 조절
        tk.Label(btn_frame, text="  크기:", bg="#eef1f6", font=("맑은 고딕", 9)).pack(side="left")
        tk.Button(btn_frame, text="A+", font=("맑은 고딕", 9, "bold"),
                  bg="#ddd", fg="#333", bd=0, padx=6, pady=6,
                  command=lambda: self._set_font_size(1)).pack(side="left", padx=2)
        tk.Button(btn_frame, text="A-", font=("맑은 고딕", 9),
                  bg="#ddd", fg="#333", bd=0, padx=6, pady=6,
                  command=lambda: self._set_font_size(-1)).pack(side="left", padx=2)

        self.report_box = scrolledtext.ScrolledText(
            root, font=("맑은 고딕", 9), wrap="word",
            bg="white", fg="#1a1a2e", bd=0, relief="flat",
            highlightthickness=1, highlightbackground="#dde2eb", highlightcolor="#3a6cf4",
            padx=18, pady=16
        )
        self.report_box.grid(row=4, column=0, columnspan=3, sticky="nsew", padx=14, pady=(6, 10))
        self.report_box.configure(state="disabled")
        setup_tags(self.report_box)

        self.status = tk.StringVar(value="파일을 선택하고 보고서 생성 버튼을 클릭하세요.")
        tk.Label(root, textvariable=self.status, bg="#eef1f6",
                 fg="#666", font=("맑은 고딕", 8)).grid(row=5, column=0, columnspan=3, pady=(0, 8))

        root.columnconfigure(1, weight=1)
        root.rowconfigure(4, weight=1)
        self._report_lines = []
        self._auto_find()

    def _auto_find(self):
        today = datetime.today().date()
        p_a = find_latest_file(FILE_DIR, 'RPA_총괄장(1차)', today)
        p_b = find_latest_file(FILE_DIR, 'RPA_총괄장(2차)', today)
        found = []
        if p_a:
            self.path_a.set(p_a)
            found.append("1차")
        if p_b:
            self.path_b.set(p_b)
            found.append("2차")
        if found:
            self.status.set(f"오늘 파일 자동 감지: {', '.join(found)} ✅  — 보고서 생성 버튼을 클릭하세요.")
        else:
            self.status.set("오늘 생성된 파일을 찾지 못했습니다. 직접 선택해주세요.")

    def _pick_a(self):
        p = filedialog.askopenfilename(title="RPA_총괄장(1차) 파일 선택",
                                       filetypes=[("CSV", "*.csv"), ("모든 파일", "*.*")],
                                       initialdir=FILE_DIR)
        if p:
            self.path_a.set(p)

    def _pick_b(self):
        p = filedialog.askopenfilename(title="RPA_총괄장(2차) 파일 선택",
                                       filetypes=[("CSV", "*.csv"), ("모든 파일", "*.*")],
                                       initialdir=FILE_DIR)
        if p:
            self.path_b.set(p)

    def _render_report(self, lines):
        box = self.report_box
        box.configure(state="normal")
        box.delete("1.0", "end")
        for tag, text in lines:
            insert_line(box, tag, text)
        box.configure(state="disabled")

    def _set_font_size(self, delta):
        self._base_size = max(7, min(18, self._base_size + delta))
        s = self._base_size
        for box in (self.report_box, self.cm_box):
            box.configure(font=("맑은 고딕", s))
            box.tag_configure("title",      font=("맑은 고딕", s + 4, "bold"))
            box.tag_configure("sub",        font=("맑은 고딕", s - 1))
            box.tag_configure("team",       font=("맑은 고딕", s + 1, "bold"))
            box.tag_configure("section",    font=("맑은 고딕", s, "bold"))
            box.tag_configure("section_box",font=("맑은 고딕", s, "bold"))
            box.tag_configure("subhead",    font=("맑은 고딕", s, "bold"))
            box.tag_configure("normal",     font=("맑은 고딕", s))
            box.tag_configure("hl_base",    font=("맑은 고딕", s))
            box.tag_configure("hl",         font=("맑은 고딕", s, "bold"))
            box.tag_configure("warn",       font=("맑은 고딕", s, "bold"))
            box.tag_configure("warn_mid",   font=("맑은 고딕", s, "bold"))
            box.tag_configure("breakdown",  font=("맑은 고딕", s - 1))
            box.tag_configure("table_hdr",  font=("맑은 고딕", s - 1))
            box.tag_configure("t_name",     font=("맑은 고딕", s, "bold"))
            box.tag_configure("t_공달",     font=("맑은 고딕", s, "bold"))
            box.tag_configure("t_공달_low", font=("맑은 고딕", s, "bold"))
            box.tag_configure("t_normal",   font=("맑은 고딕", s))
            box.tag_configure("tbl_hdr",    font=("Consolas", s, "bold"))
            box.tag_configure("tbl_row",    font=("Consolas", s))

        # 오감쇼 탭: Treeview 표 폰트/행높이 일괄 조정
        self.style.configure("Ogamsho.Treeview", font=("맑은 고딕", s), rowheight=s + 17)
        self.style.configure("Ogamsho.Treeview.Heading", font=("맑은 고딕", s, "bold"))

        self.status.set(f"글자 크기: {s}pt")
        self.og_status.set(f"글자 크기: {s}pt")
        self.cm_status.set(f"글자 크기: {s}pt")

    def _clear(self):
        self.report_box.configure(state="normal")
        self.report_box.delete("1.0", "end")
        self.report_box.configure(state="disabled")
        self._report_lines = []
        self.btn_copy.configure(state="disabled")
        self.status.set("지웠습니다.")

    def _copy(self):
        plain_lines = []
        for tag, text in self._report_lines:
            if tag == "gap":
                plain_lines.append("")
            elif tag == "partner_row":
                d = text
                brands_str = f"  ({d['brands']})" if d.get("brands") else ""
                plain_lines.append(f"    • {d['name']}{brands_str}  {d['count']}회  순{d['순달']}%  공{d['공달']}%")
            elif text:
                plain_lines.append(text)
        self.clipboard_clear()
        self.clipboard_append("\n".join(plain_lines))
        self.status.set("✅ 클립보드에 복사됐습니다.")

    def _run(self):
        if not self.path_a.get() or not self.path_b.get():
            messagebox.showwarning("파일 미선택", "1차·2차 파일을 모두 선택해주세요.")
            return
        self.btn_run.configure(state="disabled")
        self.btn_copy.configure(state="disabled")
        self.status.set("처리 중...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            df_a = pd.read_csv(self.path_a.get())
            df_b = pd.read_csv(self.path_b.get())
            lines = build_report(df_a, df_b)
            self._report_lines = lines
            self.after(0, self._render_report, lines)
            self.after(0, lambda: self.btn_copy.configure(state="normal"))
            self.after(0, lambda: self.status.set("✅ 보고서 생성 완료!"))
        except Exception as e:
            self.after(0, lambda: self.status.set("❌ 오류 발생"))
            self.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.after(0, lambda: self.btn_run.configure(state="normal"))

    # ── 오감쇼 탭 ──────────────────────────────
    def _build_ogamsho_tab(self, root):
        pad = {"padx": 12, "pady": 7}

        tk.Label(root, text="📺 오감쇼 Live방송 Daily 알리미",
                 font=("맑은 고딕", 15, "bold"), bg="#eef1f6", fg="#10254f").grid(
            row=0, column=0, columnspan=3, pady=(18, 10))

        tk.Label(root, text="RPA(1차) 파일", bg="#eef1f6", font=("맑은 고딕", 9)).grid(row=1, column=0, sticky="e", **pad)
        self.og_path = tk.StringVar()
        tk.Entry(root, textvariable=self.og_path, width=54, state="readonly").grid(row=1, column=1, sticky="ew", **pad)
        tk.Button(root, text="찾아보기", command=self._og_pick, width=8).grid(row=1, column=2, **pad)

        btn_frame = tk.Frame(root, bg="#eef1f6")
        btn_frame.grid(row=2, column=0, columnspan=3, pady=(8, 6))

        self.og_btn_run = tk.Button(btn_frame, text="  보고서 생성  ", font=("맑은 고딕", 10, "bold"),
                                    bg="#3a6cf4", fg="white", activebackground="#2756d1",
                                    bd=0, padx=16, pady=9, command=self._og_run)
        self.og_btn_run.pack(side="left", padx=6)

        self.og_btn_copy = tk.Button(btn_frame, text="  클립보드 복사  ", font=("맑은 고딕", 10),
                                     bg="#22a85a", fg="white", activebackground="#1a8247",
                                     bd=0, padx=16, pady=9, command=self._og_copy, state="disabled")
        self.og_btn_copy.pack(side="left", padx=6)

        self.og_btn_clear = tk.Button(btn_frame, text="  지우기  ", font=("맑은 고딕", 10),
                                      bg="#888", fg="white", activebackground="#666",
                                      bd=0, padx=16, pady=9, command=self._og_clear)
        self.og_btn_clear.pack(side="left", padx=6)

        tk.Label(btn_frame, text="  크기:", bg="#eef1f6", font=("맑은 고딕", 9)).pack(side="left")
        tk.Button(btn_frame, text="A+", font=("맑은 고딕", 9, "bold"),
                  bg="#ddd", fg="#333", bd=0, padx=6, pady=6,
                  command=lambda: self._set_font_size(1)).pack(side="left", padx=2)
        tk.Button(btn_frame, text="A-", font=("맑은 고딕", 9),
                  bg="#ddd", fg="#333", bd=0, padx=6, pady=6,
                  command=lambda: self._set_font_size(-1)).pack(side="left", padx=2)

        # ── Treeview 표 스타일 (헤더: 남색 배경/흰 글씨, 본문: 줄무늬) ──
        self.style.configure("Ogamsho.Treeview", font=("맑은 고딕", 9), rowheight=26,
                              background="white", fieldbackground="white", borderwidth=0)
        self.style.configure("Ogamsho.Treeview.Heading", font=("맑은 고딕", 9, "bold"),
                              background="#26314a", foreground="white", relief="flat")
        self.style.map("Ogamsho.Treeview.Heading", background=[("active", "#26314a")])

        # ── 스크롤 가능한 표 대시보드 ──
        canvas = tk.Canvas(root, bg="#eef1f6", highlightthickness=0)
        vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg="#eef1f6")
        inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)

        def _on_inner_configure(_e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(e):
            canvas.itemconfig(inner_id, width=e.width)

        inner.bind("<Configure>", _on_inner_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        canvas.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=(14, 0), pady=(6, 10))
        vsb.grid(row=3, column=2, sticky="ns", pady=(6, 10))
        root.columnconfigure(1, weight=1)
        root.rowconfigure(3, weight=1)

        self.og_title_label = tk.Label(inner, text="", font=("맑은 고딕", 13, "bold"),
                                        bg="#eef1f6", fg="#10254f", anchor="w")
        self.og_title_label.pack(fill="x", padx=14, pady=(4, 0))
        self.og_summary_label = tk.Label(inner, text="", font=("맑은 고딕", 9),
                                          bg="#eef1f6", fg="#5a6472", anchor="w", justify="left")
        self.og_summary_label.pack(fill="x", padx=14, pady=(2, 8))

        self.og_tv_daily, self._og_daily_frame = self._make_table(
            inner, "📌 당일 방송 실적  (상품명 더블클릭 → 상품 페이지 열기)",
            ["브랜드", "상품", "시간", "노출분", "전체달성률", "생방달성률"],
            [90, 280, 100, 60, 90, 90])
        self._og_daily_urls = {}
        self.og_tv_daily.bind("<Double-1>", self._og_open_daily_link)

        self.og_tv_cum, _ = self._make_table(
            inner, "📌 누적 실적",
            ["구분", "횟수", "매출", "공헌", "광고비", "순달", "공달"],
            [110, 60, 90, 90, 90, 70, 70])

        self.og_tv_yoy, _ = self._make_table(
            inner, "📌 연도별 비교 (동기간 기준)",
            ["구분", "전년 매출", "금년 매출", "증감", "전년 달성률", "금년 달성률"],
            [120, 90, 90, 70, 120, 120])

        self.og_tv_monthly, _ = self._make_table(
            inner, "📌 월별 운영 추이  (가장 최근월 강조, 연간 총계 포함)",
            ["기준년월", "횟수", "매출", "공헌", "순달", "공달"],
            [110, 60, 90, 90, 70, 70])

        self.og_tv_newprod, _ = self._make_table(
            inner, "📌 신상품/기존상품 비교 (연도별)",
            ["연도", "구분", "PGM수", "비중(%)", "순달", "공달"],
            [90, 130, 60, 70, 70, 70])

        # 카테고리 표 + 파이차트를 좌우로 배치
        cat_row = tk.Frame(inner, bg="#eef1f6")
        cat_row.pack(fill="x", padx=14, pady=(0, 10))
        cat_row.columnconfigure(0, weight=1)
        cat_row.columnconfigure(1, weight=0)

        cat_card = tk.Frame(cat_row, bg="white", highlightthickness=1, highlightbackground="#dde2eb")
        cat_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.og_cat_label = tk.Label(cat_card, text="📌 카테고리별 실적", font=("맑은 고딕", 10, "bold"),
                                      bg="white", fg="#1a3a6e", anchor="w")
        self.og_cat_label.pack(fill="x", padx=12, pady=(10, 4))
        cat_cols = ["카테고리", "횟수", "비중", "매출", "공헌", "순달", "공달", "순달(생)", "공달(생)"]
        self.og_tv_cat = ttk.Treeview(cat_card, columns=cat_cols,
                                       show="headings", style="Ogamsho.Treeview", height=1)
        for col, w in zip(cat_cols, [90, 55, 55, 80, 80, 55, 55, 65, 65]):
            self.og_tv_cat.heading(col, text=col)
            self.og_tv_cat.column(col, width=w, anchor="center")
        self.og_tv_cat.pack(fill="x", padx=12, pady=(0, 12))
        self.og_tv_cat.tag_configure("hl_strong", background="#fff4cc", font=("맑은 고딕", 9, "bold"))
        self.og_tv_cat.tag_configure("hl_total",  background="#e3e7f0", font=("맑은 고딕", 9, "bold"))

        self.og_chart_card = tk.Frame(cat_row, bg="white", highlightthickness=1, highlightbackground="#dde2eb")
        self.og_chart_card.grid(row=0, column=1, sticky="ns")
        self.og_chart_widget = None

        self.og_status = tk.StringVar(
            value="파일을 선택하고 보고서 생성 버튼을 클릭하세요. (생성 시 분석 엑셀도 Downloads에 자동 저장됩니다)")
        tk.Label(root, textvariable=self.og_status, bg="#eef1f6",
                 fg="#666", font=("맑은 고딕", 8), wraplength=680, justify="left").grid(
            row=4, column=0, columnspan=3, pady=(0, 8))

        self._og_report_lines = []
        self._og_data = None
        self._og_auto_find()

    def _make_table(self, parent, title, columns, widths):
        frame = tk.Frame(parent, bg="white", highlightthickness=1, highlightbackground="#dde2eb")
        frame.pack(fill="x", padx=14, pady=(0, 10))
        tk.Label(frame, text=title, font=("맑은 고딕", 10, "bold"), bg="white", fg="#1a3a6e",
                 anchor="w").pack(fill="x", padx=12, pady=(10, 4))
        tv = ttk.Treeview(frame, columns=columns, show="headings", style="Ogamsho.Treeview", height=1)
        for col, w in zip(columns, widths):
            tv.heading(col, text=col)
            tv.column(col, width=w, anchor="center")
        tv.pack(fill="x", padx=12, pady=(0, 12))
        tv.tag_configure("hl_good",   foreground="#1f9d55", font=("맑은 고딕", 9, "bold"))
        tv.tag_configure("hl_bad",    foreground="#d6311a", font=("맑은 고딕", 9, "bold"))
        tv.tag_configure("hl_strong", background="#fff4cc", font=("맑은 고딕", 9, "bold"))
        tv.tag_configure("hl_total",  background="#e3e7f0", font=("맑은 고딕", 9, "bold"))
        tv.tag_configure("hl_new",    foreground="#0860d6", font=("맑은 고딕", 9, "bold"))
        tv.tag_configure("link_row",  foreground="#0860d6")
        return tv, frame

    def _og_open_daily_link(self, _event):
        sel = self.og_tv_daily.selection()
        if not sel:
            return
        url = self._og_daily_urls.get(sel[0])
        if url:
            webbrowser.open(url)

    def _og_auto_find(self):
        today = datetime.today().date()
        p = find_latest_file(FILE_DIR, 'RPA_총괄장(1차)', today)
        if p:
            self.og_path.set(p)
            self.og_status.set("오늘 파일 자동 감지 ✅  — 보고서 생성 버튼을 클릭하세요.")
            if hasattr(self, 'cm_path'):
                self.cm_path.set(p)
        else:
            self.og_status.set("오늘 생성된 RPA(1차) 파일을 찾지 못했습니다. 직접 선택해주세요.")

    def _og_pick(self):
        p = filedialog.askopenfilename(title="RPA_총괄장(1차) 파일 선택",
                                       filetypes=[("CSV", "*.csv"), ("모든 파일", "*.*")],
                                       initialdir=FILE_DIR)
        if p:
            self.og_path.set(p)

    def _og_fill(self, tv, rows, height=None):
        tv.delete(*tv.get_children())
        for iid, (values, tags) in enumerate(rows):
            tv.insert("", "end", iid=str(iid), values=values, tags=tags)
        tv.configure(height=max(1, height if height is not None else len(rows)))

    def _og_render_tables(self, data):
        if data is None:
            self.og_title_label.configure(text="데이터 없음")
            self.og_summary_label.configure(text="선택한 파일에서 오감쇼 방송 데이터를 찾지 못했습니다.")
            self.og_cat_label.configure(text="📌 카테고리별 실적")
            for tv in (self.og_tv_daily, self.og_tv_cum, self.og_tv_yoy,
                       self.og_tv_monthly, self.og_tv_newprod, self.og_tv_cat):
                tv.delete(*tv.get_children())
            if self.og_chart_widget:
                self.og_chart_widget.destroy()
                self.og_chart_widget = None
            return

        self.og_title_label.configure(text=f"📺 {data['title']}")
        ds = data['daily_summary']
        self.og_summary_label.configure(
            text=f"매출 {ds['sales']} / 공헌 {ds['profit']} / 광고비 {ds['ad']}   |   "
                 f"전체 달성률 순{ds['sn']}%-공{ds['sc']}%   |   생방송 달성률 순{ds['ln']}%-공{ds['lc']}%")

        # 당일 방송 실적
        self._og_daily_urls = {}
        daily_disp = []
        for r in data['daily_rows']:
            daily_disp.append(((r['brand'], r['product'], r['time'], r['expo'],
                                 f"순{r['sn']}%-공{r['sc']}%", f"순{r['ln']}%-공{r['lc']}%"),
                                ("link_row",) if r['url'] else ()))
        self._og_fill(self.og_tv_daily, daily_disp)
        for iid, r in enumerate(data['daily_rows']):
            if r['url']:
                self._og_daily_urls[str(iid)] = r['url']

        # 누적 실적
        cum_disp = []
        for r in data['cumulative']:
            tags = ("hl_strong",) if r['label'].endswith("전체") and "년" in r['label'] else ()
            cum_disp.append(((r['label'], r['count'], r['sales'], r['profit'], r['ad'],
                               f"{r['sn']}%", f"{r['sc']}%"), tags))
        self._og_fill(self.og_tv_cum, cum_disp)

        # 연도별 비교
        yoy_disp = []
        for r in data['yoy']:
            delta_str = "-" if r['delta'] is None else f"{'+' if r['delta'] >= 0 else ''}{r['delta']}%"
            tags = ("hl_good",) if (r['delta'] or 0) >= 0 else ("hl_bad",)
            yoy_disp.append(((r['label'], r['sales_prev'], r['sales_cur'], delta_str,
                               f"순{r['sn_prev']}%-공{r['sc_prev']}%", f"순{r['sn_cur']}%-공{r['sc_cur']}%"), tags))
        self._og_fill(self.og_tv_yoy, yoy_disp)

        # 월별 운영 추이 (가장 최근월 강조 + 연간 총계)
        mon_disp = []
        for r in data['monthly']:
            if r.get('is_total'):
                tags = ("hl_total",)
            elif r['label'] == data['highlight_ym']:
                tags = ("hl_strong",)
            else:
                tags = ()
            mon_disp.append(((r['label'], r['count'], r['sales'], r['profit'],
                               f"{r['sn']}%", f"{r['sc']}%"), tags))
        self._og_fill(self.og_tv_monthly, mon_disp)

        # 신상품/기존상품 비교 (연도별) — 같은 연도가 반복되면 빈칸으로 둬서 셀병합처럼 보이게 함
        np_disp = []
        prev_year = None
        for r in data.get('newproduct', []):
            year_disp = r['year'] if r['year'] != prev_year else ""
            prev_year = r['year']
            if r.get('is_total'):
                tags = ("hl_total",)
            elif r.get('is_new'):
                tags = ("hl_new",)
            else:
                tags = ()
            np_disp.append(((year_disp, r['label'], r['count'], f"{r['weight']}%",
                              f"{r['sn']}%", f"{r['sc']}%"), tags))
        self._og_fill(self.og_tv_newprod, np_disp)

        # 카테고리별 실적 (Total 행 포함)
        self.og_cat_label.configure(text=f"📌 {data['year']}년 카테고리별 실적")
        self.og_tv_cat.delete(*self.og_tv_cat.get_children())
        for iid, r in enumerate(data['category']):
            if r.get('is_total'):
                tags = ("hl_total",)
            elif iid == 0:
                tags = ("hl_strong",)
            else:
                tags = ()
            self.og_tv_cat.insert("", "end", iid=str(iid),
                                   values=(r['label'], r['count'], f"{r['weight']}%",
                                           r['sales'], r['profit'], f"{r['sn']}%", f"{r['sc']}%",
                                           f"{r['sn_live']}%", f"{r['sc_live']}%"),
                                   tags=tags)
        self.og_tv_cat.configure(height=max(1, len(data['category'])))

        # 카테고리별 편성비중 파이차트 (Total 행은 제외)
        if self.og_chart_widget:
            self.og_chart_widget.destroy()
            self.og_chart_widget = None
        chart_rows = [r for r in data['category'] if not r.get('is_total')]
        if chart_rows:
            self.og_chart_widget = self._make_pie_chart(self.og_chart_card, chart_rows)
            self.og_chart_widget.pack(padx=12, pady=12)

    def _make_pie_chart(self, parent, category_rows):
        fig = Figure(figsize=(3.6, 3.2), dpi=90, facecolor="white")
        ax = fig.add_subplot(111)
        labels = [r['label'] for r in category_rows]
        sizes  = [r['weight_raw'] for r in category_rows]
        colors = ["#3a6cf4", "#22a85a", "#f5a623", "#d6311a", "#7030a0", "#17a2b8", "#6b7280", "#e91e8c"]
        ax.pie(sizes, labels=labels, autopct="%1.0f%%", startangle=90,
               colors=[colors[i % len(colors)] for i in range(len(labels))],
               textprops={"fontsize": 8})
        ax.set_title("카테고리별 편성비중", fontsize=10, fontweight="bold")
        canvas_widget = FigureCanvasTkAgg(fig, master=parent)
        canvas_widget.draw()
        return canvas_widget.get_tk_widget()

    def _og_clear(self):
        self._og_render_tables(None)
        self._og_report_lines = []
        self.og_btn_copy.configure(state="disabled")
        self.og_status.set("지웠습니다.")

    def _og_copy(self):
        plain_lines = []
        for tag, text in self._og_report_lines:
            if tag == "gap":
                plain_lines.append("")
            elif tag in ("ytable_hdr", "ytable_row"):
                plain_lines.append("\t".join(text))
            elif tag == "product_line":
                d = text
                suffix = f"  ({d['url']})" if d.get("url") else ""
                plain_lines.append(f"  • {d['brand']}  {d['product']}{suffix}")
            else:
                plain_lines.append(text)
        self.clipboard_clear()
        self.clipboard_append("\n".join(plain_lines))
        self.og_status.set("✅ 클립보드에 복사됐습니다.")

    def _og_run(self):
        if not self.og_path.get():
            messagebox.showwarning("파일 미선택", "RPA(1차) 파일을 선택해주세요.")
            return
        self.og_btn_run.configure(state="disabled")
        self.og_btn_copy.configure(state="disabled")
        self.og_status.set("처리 중...")
        threading.Thread(target=self._og_worker, daemon=True).start()

    def _og_worker(self):
        try:
            df = load_ogamsho_df(self.og_path.get())
            lines = build_ogamsho_report(df)
            data  = build_ogamsho_tables(df)
            self._og_report_lines = lines
            self._og_data = df
            self.after(0, self._og_render_tables, data)
            self.after(0, lambda: self.og_btn_copy.configure(state="normal"))
            if not df.empty:
                outpath = create_ogamsho_excel(df, lines)
                self.after(0, lambda: self.og_status.set(f"✅ 보고서 생성 완료! 분석 엑셀 저장: {outpath}"))
            else:
                self.after(0, lambda: self.og_status.set("⚠ 오감쇼/황정민쇼 데이터가 없습니다."))
        except Exception as e:
            self.after(0, lambda: self.og_status.set("❌ 오류 발생"))
            self.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.after(0, lambda: self.og_btn_run.configure(state="normal"))

    # ── 보고멘트 탭 ────────────────────────────
    def _build_comment_tab(self, root):
        pad = {"padx": 12, "pady": 7}

        tk.Label(root, text="📝 오감쇼 보고멘트 (최신 방송일 기준 초안)",
                 font=("맑은 고딕", 15, "bold"), bg="#eef1f6", fg="#10254f").grid(
            row=0, column=0, columnspan=3, pady=(18, 10))

        tk.Label(root, text="RPA(1차) 파일", bg="#eef1f6", font=("맑은 고딕", 9)).grid(row=1, column=0, sticky="e", **pad)
        self.cm_path = tk.StringVar()
        if getattr(self, 'og_path', None) and self.og_path.get():
            self.cm_path.set(self.og_path.get())
        tk.Entry(root, textvariable=self.cm_path, width=54, state="readonly").grid(row=1, column=1, sticky="ew", **pad)
        tk.Button(root, text="찾아보기", command=self._cm_pick, width=8).grid(row=1, column=2, **pad)

        btn_frame = tk.Frame(root, bg="#eef1f6")
        btn_frame.grid(row=2, column=0, columnspan=3, pady=(8, 6))

        self.cm_btn_run = tk.Button(btn_frame, text="  보고멘트 생성  ", font=("맑은 고딕", 10, "bold"),
                                    bg="#3a6cf4", fg="white", activebackground="#2756d1",
                                    bd=0, padx=16, pady=9, command=self._cm_run)
        self.cm_btn_run.pack(side="left", padx=6)

        self.cm_btn_copy = tk.Button(btn_frame, text="  클립보드 복사  ", font=("맑은 고딕", 10),
                                     bg="#22a85a", fg="white", activebackground="#1a8247",
                                     bd=0, padx=16, pady=9, command=self._cm_copy, state="disabled")
        self.cm_btn_copy.pack(side="left", padx=6)

        self.cm_btn_clear = tk.Button(btn_frame, text="  지우기  ", font=("맑은 고딕", 10),
                                      bg="#888", fg="white", activebackground="#666",
                                      bd=0, padx=16, pady=9, command=self._cm_clear)
        self.cm_btn_clear.pack(side="left", padx=6)

        tk.Label(btn_frame, text="  크기:", bg="#eef1f6", font=("맑은 고딕", 9)).pack(side="left")
        tk.Button(btn_frame, text="A+", font=("맑은 고딕", 9, "bold"),
                  bg="#ddd", fg="#333", bd=0, padx=6, pady=6,
                  command=lambda: self._set_font_size(1)).pack(side="left", padx=2)
        tk.Button(btn_frame, text="A-", font=("맑은 고딕", 9),
                  bg="#ddd", fg="#333", bd=0, padx=6, pady=6,
                  command=lambda: self._set_font_size(-1)).pack(side="left", padx=2)

        tk.Label(root, text="※ 수치(매출/공헌/달성률/누적실적)는 자동 계산되며, '리뷰 내용을 입력해주세요' 부분은 직접 작성해주세요.",
                 bg="#eef1f6", fg="#888", font=("맑은 고딕", 8)).grid(
            row=3, column=0, columnspan=3, sticky="w", padx=14)

        self.cm_box = scrolledtext.ScrolledText(
            root, font=("맑은 고딕", 9), wrap="word",
            bg="white", fg="#1a1a2e", bd=0, relief="flat",
            highlightthickness=1, highlightbackground="#dde2eb", highlightcolor="#3a6cf4",
            padx=18, pady=16
        )
        # 직접 수정(리뷰 작성)할 수 있도록 편집 가능 상태로 둠
        self.cm_box.grid(row=4, column=0, columnspan=3, sticky="nsew", padx=14, pady=(6, 10))

        self.cm_status = tk.StringVar(value="파일을 선택하고 보고멘트 생성 버튼을 클릭하세요.")
        tk.Label(root, textvariable=self.cm_status, bg="#eef1f6",
                 fg="#666", font=("맑은 고딕", 8)).grid(row=5, column=0, columnspan=3, pady=(0, 8))

        root.columnconfigure(1, weight=1)
        root.rowconfigure(4, weight=1)

    def _cm_pick(self):
        p = filedialog.askopenfilename(title="RPA_총괄장(1차) 파일 선택",
                                       filetypes=[("CSV", "*.csv"), ("모든 파일", "*.*")],
                                       initialdir=FILE_DIR)
        if p:
            self.cm_path.set(p)

    def _cm_clear(self):
        self.cm_box.delete("1.0", "end")
        self.cm_btn_copy.configure(state="disabled")
        self.cm_status.set("지웠습니다.")

    def _cm_copy(self):
        self.clipboard_clear()
        self.clipboard_append(self.cm_box.get("1.0", "end-1c"))
        self.cm_status.set("✅ 클립보드에 복사됐습니다.")

    def _cm_run(self):
        if not self.cm_path.get():
            messagebox.showwarning("파일 미선택", "RPA(1차) 파일을 선택해주세요.")
            return
        self.cm_btn_run.configure(state="disabled")
        self.cm_btn_copy.configure(state="disabled")
        self.cm_status.set("처리 중...")
        threading.Thread(target=self._cm_worker, daemon=True).start()

    def _cm_worker(self):
        try:
            df = load_ogamsho_df(self.cm_path.get())
            text = build_ogamsho_comment(df)
            self.after(0, self._cm_render, text)
            self.after(0, lambda: self.cm_btn_copy.configure(state="normal"))
            self.after(0, lambda: self.cm_status.set("✅ 보고멘트 초안 생성 완료! 리뷰 부분을 채워주세요."))
        except Exception as e:
            self.after(0, lambda: self.cm_status.set("❌ 오류 발생"))
            self.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.after(0, lambda: self.cm_btn_run.configure(state="normal"))

    def _cm_render(self, text):
        self.cm_box.delete("1.0", "end")
        self.cm_box.insert("1.0", text)


if __name__ == "__main__":
    app = App()
    app.mainloop()